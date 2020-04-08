from openpyxl import load_workbook

import helpers


class UAVData:
    def __init__(self, excel: str, sheet_n: int = 0, n_head_row: int = 0):
        print("\tMessage: UAV data processing...")

        book = load_workbook(excel, read_only=True)
        sheet = book.worksheets[sheet_n]

        self.n_head_row = n_head_row
        self.n_row = sheet.max_row - n_head_row
        self.n_column = sheet.max_column
        self.properties = tuple(sheet.cell(1, i + 1).value for i in range(self.n_column))
        if n_head_row:
            self.units = tuple(sheet.cell(2, i + 1).value for i in range(self.n_column))

        self.data = []
        for i in range(self.n_row - self.n_head_row):
            self.data.append(tuple(sheet.cell(i + self.n_head_row + 1, j + 1).value for j in range(self.n_column)))

        print("\tMessage: UAV data ready")


class UAVGeometry:
    def __init__(self, w: float, h: float, l: float, unit: str = 'm'):
        w = .1 if w is None else w
        h = .1 if h is None else h
        l = .1 if l is None else l
        if w < 0 or h < 0 or l < 0:
            print("\tError in <UAVGeometry>: sizes must be positive!")
            raise ValueError()

        self.unit, self.mul = unit, 1
        if unit == 'cm':
            self.mul = 0.01
        elif unit == 'mm':
            self.mul = 0.001
        self.w, self.h, self.len = w * self.mul, h * self.mul, l * self.mul

    def __str__(self):
        return f"(width x height x length): {self.w} x {self.h} x {self.len} [m]"


class UAV:
    def __init__(self, properties: tuple, units: tuple, data: tuple):
        if len(properties) != len(units) != len(data):
            print("\tError in <UAV>: lengths of input arguments must be equal to each other!")
            raise ValueError

        self.units = self.get_units(properties, units)
        self.data = self.get_data(properties, data)

        self.model = self.data['model']
        self.country = self.data['country']
        self.endurance = self.data['endurance']
        self.max_range = self.data['range']
        self.payload = self.data['payload']
        self.max_speed = self.data['max speed']
        self.ceiling = self.data['ceiling']
        self.takeoff_weight = self.data['max takeoff weight']
        self.geometry = UAVGeometry(self.data['width'], self.data['height'], self.data['length'], unit='m')
        self.temp_low = self.data['temp. low']
        self.temp_up = self.data['temp. up']
        self.url_source = self.data['source']

        self.freq_str = self.get_frequency_list(self.data['frequency']) if self.data['frequency'] is not None else None
        self.freq = None
        if self.freq_str is not None:
            self.init_frequency_values()

    def __str__(self):
        return f"UAV model {self.model}:" \
               f"\n - country: {self.country}" \
               f"\n - endurance: {self.endurance} [{self.units['endurance']}]" \
               f"\n - max range: {self.max_range} [{self.units['range']}]" \
               f"\n - payload: {self.payload} [{self.units['payload']}]" \
               f"\n - max speed: {self.max_speed} [{self.units['max speed']}]" \
               f"\n - ceiling: {self.ceiling} [{self.units['ceiling']}]" \
               f"\n - max takeoff weight: {self.takeoff_weight} [{self.units['max takeoff weight']}]" \
               f"\n - geometry {self.geometry}" \
               f"\n - frequencies ranges names: {self.freq_str}" \
               f"\n - frequencies values: {self.freq} [{self.units['frequency']}]" \
               f"\n - operating temperatures: from {self.temp_low} to {self.temp_up} [{self.units['temp. low']}]" \
               f"\n - info source: {self.url_source}"

    @staticmethod
    def get_units(keys: tuple, data: tuple) -> dict:
        return helpers.items_to_float({keys[i].lower(): data[i] for i in range(len(keys))})

    @staticmethod
    def get_data(keys: tuple, data: tuple) -> dict:
        return {keys[i].lower(): data[i] for i in range(len(keys))}

    @staticmethod
    def get_frequency_list(freqs: str, delim: str = ', ') -> list:
        return freqs.split(delim)

    def init_frequency_values(self):
        self.freq = []
        for f in self.freq_str:
            if f.lower() == 'wi-fi':
                self.freq.append(2400)            # GHz
            elif f.lower() == 'gps l1':
                self.freq.append(1500)
            elif f.lower() == 'gps l2-l5':
                self.freq.append(1200)
            elif f.lower() == '5.2g':
                self.freq.append(5200)
            elif f.lower() == '5.8g':
                self.freq.append(5800)
            elif helpers.is_digit(f):
                self.freq.append(float(f))
            else:
                print(f"\tWarning in <UAV>: unknown frequency range '{f}'!")
                self.freq.append(None)
